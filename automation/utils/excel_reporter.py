import os
import logging
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger("ExcelReporter")

def generate_excel_reports(test_results, metrics, output_dir="reports/excel"):
    os.makedirs(output_dir, exist_ok=True)
    main_report_path = os.path.join(output_dir, "Automation_Test_Report.xlsx")
    passed_report_path = os.path.join(output_dir, "Passed_Test_Cases.xlsx")
    failed_report_path = os.path.join(output_dir, "Failed_Test_Cases.xlsx")
    summary_report_path = os.path.join(output_dir, "Execution_Summary.xlsx")

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Styles
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    passed_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    failed_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    skipped_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    bold_font = Font(name="Arial", size=10, bold=True)
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    headers = ["Test ID", "Module", "Test Name", "Priority", "Status", "Execution Time (s)", "Failure Reason"]

    # --- Sheet 1: Executed Test Cases ---
    ws_all = wb.create_sheet("Executed Test Cases")
    ws_all.append(headers)
    for res in test_results:
        ws_all.append([
            res["test_id"], res["module"], res["test_name"],
            res["priority"], res["status"], res["duration"],
            res.get("failure_reason", "")
        ])

    # --- Sheet 2: Passed Tests ---
    ws_passed = wb.create_sheet("Passed Tests")
    ws_passed.append(headers)
    for res in test_results:
        if res["status"] == "PASSED":
            ws_passed.append([
                res["test_id"], res["module"], res["test_name"],
                res["priority"], res["status"], res["duration"], ""
            ])

    # --- Sheet 3: Failed Tests ---
    ws_failed = wb.create_sheet("Failed Tests")
    ws_failed.append(headers)
    for res in test_results:
        if res["status"] == "FAILED":
            ws_failed.append([
                res["test_id"], res["module"], res["test_name"],
                res["priority"], res["status"], res["duration"],
                res.get("failure_reason", "Assertion Failed")
            ])

    # --- Sheet 4: Skipped Tests ---
    ws_skipped = wb.create_sheet("Skipped Tests")
    ws_skipped.append(headers)
    for res in test_results:
        if res["status"] == "SKIPPED":
            ws_skipped.append([
                res["test_id"], res["module"], res["test_name"],
                res["priority"], res["status"], res["duration"],
                res.get("failure_reason", "Feature Disabled")
            ])

    # --- Sheet 5: Execution Metrics ---
    ws_metrics = wb.create_sheet("Execution Metrics")
    ws_metrics.append(["Metric", "Value"])
    ws_metrics.append(["Total Test Cases", metrics["total"]])
    ws_metrics.append(["Executed", metrics["executed"]])
    ws_metrics.append(["Passed", metrics["passed"]])
    ws_metrics.append(["Failed", metrics["failed"]])
    ws_metrics.append(["Skipped", metrics["skipped"]])
    ws_metrics.append(["Pass Percentage", f"{metrics['pass_rate']}%"])
    ws_metrics.append(["Total Execution Duration", f"{metrics['duration']}s"])

    # --- Sheet 6: Defect Summary ---
    ws_defects = wb.create_sheet("Defect Summary")
    ws_defects.append(["Defect ID", "Test Case ID", "Module", "Severity", "Failure Reason"])
    defect_id = 101
    for res in test_results:
        if res["status"] == "FAILED":
            ws_defects.append([
                f"DEF_{defect_id}", res["test_id"], res["module"],
                "High" if res["priority"] in ["P0", "P1"] else "Medium",
                res.get("failure_reason", "Validation Error")
            ])
            defect_id += 1

    # --- Sheet 7: Pass Rate Summary ---
    ws_passrate = wb.create_sheet("Pass Rate Summary")
    ws_passrate.append(["Module", "Total", "Passed", "Failed", "Skipped", "Pass Rate (%)"])
    module_stats = {}
    for res in test_results:
        m = res["module"]
        if m not in module_stats:
            module_stats[m] = {"total": 0, "passed": 0, "failed": 0, "skipped": 0}
        module_stats[m]["total"] += 1
        if res["status"] == "PASSED":
            module_stats[m]["passed"] += 1
        elif res["status"] == "FAILED":
            module_stats[m]["failed"] += 1
        else:
            module_stats[m]["skipped"] += 1

    for mod, s in module_stats.items():
        pr = round((s["passed"] / s["total"]) * 100, 2) if s["total"] > 0 else 0
        ws_passrate.append([mod, s["total"], s["passed"], s["failed"], s["skipped"], f"{pr}%"])

    # Apply styling & auto-fit columns for all sheets
    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin_border
                if cell.value == "PASSED":
                    cell.fill = passed_fill
                elif cell.value == "FAILED":
                    cell.fill = failed_fill
                elif cell.value == "SKIPPED":
                    cell.fill = skipped_fill

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    wb.save(main_report_path)
    logger.info(f"Saved main report to {main_report_path}")

    # Export individual filtered workbooks
    wb_p = openpyxl.Workbook()
    ws_p = wb_p.active
    ws_p.title = "Passed Tests"
    ws_p.append(headers)
    for res in test_results:
        if res["status"] == "PASSED":
            ws_p.append([res["test_id"], res["module"], res["test_name"], res["priority"], res["status"], res["duration"], ""])
    wb_p.save(passed_report_path)

    wb_f = openpyxl.Workbook()
    ws_f = wb_f.active
    ws_f.title = "Failed Tests"
    ws_f.append(headers)
    for res in test_results:
        if res["status"] == "FAILED":
            ws_f.append([res["test_id"], res["module"], res["test_name"], res["priority"], res["status"], res["duration"], res.get("failure_reason", "")])
    wb_f.save(failed_report_path)

    wb_s = openpyxl.Workbook()
    ws_s = wb_s.active
    ws_s.title = "Summary"
    ws_s.append(["Metric", "Value"])
    for k, v in metrics.items():
        ws_s.append([k, str(v)])
    wb_s.save(summary_report_path)
